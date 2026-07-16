"""
coretax_export.py
=================
#5 dari roadmap: Export PPh 23 bulanan dari ap_invoice ke Excel
yang siap di-copy ke Converter XML Coretax resmi DJP (Bupot Unifikasi).

CATATAN PENTING:
  - Coretax (2025+) pakai skema import XML. DJP menyediakan template
    Excel converter yang diisi → di-export ke XML → upload ke Coretax.
  - Script ini generate Excel dengan kolom STANDAR Bupot PPh 23 supaya
    datanya tinggal di-copy ke converter resmi DJP (versi terbaru).
  - Selalu cross-check kolom dengan converter DJP terbaru karena
    skema bisa berubah. Sheet "PANDUAN" berisi mapping kolomnya.

Output: 3 sheet
  1. DATA       — baris bupot siap copy (1 baris per invoice)
  2. PANDUAN    — penjelasan tiap kolom + catatan validasi
  3. REKAP      — ringkasan: jumlah bupot, total DPP, total PPh

Cara jalankan:
  python coretax_export.py --month 2026-03
  python coretax_export.py --month 2026-03 --entity-id <uuid>
  python coretax_export.py --month 2026-03 --output bupot_maret.xlsx
"""

import os
import sys
import argparse
from decimal import Decimal
from datetime import date, datetime

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    f"postgresql://{os.getenv('DB_USER','postgres')}:{os.getenv('DB_PASSWORD','')}@"
    f"{os.getenv('DB_HOST','localhost')}:{os.getenv('DB_PORT','5432')}/"
    f"{os.getenv('DB_NAME','accounting_db')}"
)

# Kode objek pajak PPh 23 (umum). Sesuaikan dengan master data DJP terbaru.
# Mapping treatment → kode objek pajak default
PPH23_OBJEK_DEFAULT = "24-104-01"   # Jasa lain (PMK 141) — paling umum
PPH23_OBJEK_MAP = {
    "jasa_manajemen":  "24-104-06",   # Jasa manajemen
    "jasa_teknik":     "24-104-02",   # Jasa teknik
    "jasa_konsultan":  "24-104-05",   # Jasa konsultan
    "sewa":            "24-100-03",   # Sewa selain tanah/bangunan
}

# Status invoice yang dimasukkan ke bupot (yang sudah pasti dipotong)
EXPORT_STATUSES = ("approved", "partial", "paid")

# Styling
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT  = Font(name="Arial", bold=True, size=13, color="1F4E78")
NOTE_FONT   = Font(name="Arial", italic=True, size=9, color="808080")
CELL_FONT   = Font(name="Arial", size=10)
THIN        = Side(style="thin", color="D0D0D0")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─────────────────────────────────────────────────────────────────────────────
# FETCH
# ─────────────────────────────────────────────────────────────────────────────

def get_bupot_data(session, year, month, entity_id=None):
    """Ambil ap_invoice yang kena PPh 23 untuk periode tertentu."""
    entity_filter = "AND ai.entity_id = :entity_id" if entity_id else ""
    params = {"year": year, "month": month, "statuses": tuple(EXPORT_STATUSES)}
    if entity_id:
        params["entity_id"] = str(entity_id)

    rows = session.execute(
        text(f"""
            SELECT
                ai.id,
                ai.invoice_no,
                ai.invoice_date,
                ai.subtotal,
                ai.pph_type,
                ai.pph_rate,
                ai.pph_amount,
                v.vendor_name,
                v.npwp,
                v.vendor_code,
                v.vendor_category,
                v.has_skb
            FROM ap_invoice ai
            JOIN vendor v ON v.id = ai.vendor_id
            WHERE EXTRACT(YEAR  FROM ai.invoice_date) = :year
              AND EXTRACT(MONTH FROM ai.invoice_date) = :month
              AND ai.status IN :statuses
              AND ai.pph_amount > 0
              AND ai.pph_type LIKE 'PPh23%'
              {entity_filter}
            ORDER BY ai.invoice_date ASC, v.vendor_name ASC
        """),
        params
    ).fetchall()
    return [dict(r._mapping) for r in rows]

# ─────────────────────────────────────────────────────────────────────────────
# BUILD EXCEL
# ─────────────────────────────────────────────────────────────────────────────

# Kolom DATA — disusun mengikuti urutan umum converter Bupot PPh 23 DJP.
# Sesuaikan label/urutan dengan converter resmi versi terbaru saat copy.
DATA_COLUMNS = [
    ("No",                 8),
    ("Masa Pajak",         12),
    ("Tahun Pajak",        12),
    ("NPWP/NIK Pihak Dipotong", 24),
    ("Nama Pihak Dipotong", 30),
    ("Kode Objek Pajak",   16),
    ("DPP (Rp)",           16),
    ("Tarif (%)",          10),
    ("PPh Dipotong (Rp)",  18),
    ("No Dokumen Referensi", 20),
    ("Tanggal Dokumen",    16),
    ("Keterangan",         28),
]

def normalize_npwp(npwp):
    """Bersihkan NPWP: hapus titik/strip, jadikan 15/16 digit."""
    if not npwp:
        return ""
    return "".join(c for c in str(npwp) if c.isdigit())


def build_workbook(rows, year, month, entity_id):
    wb = Workbook()

    # ============ SHEET 1: DATA ============
    ws = wb.active
    ws.title = "DATA"

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"DAFTAR BUKTI POTONG PPh PASAL 23 — MASA {month:02d}/{year}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:L2")
    ws["A2"] = ("Copy baris data (mulai baris 5) ke Converter XML Bupot Unifikasi resmi DJP, "
                "lalu export ke XML untuk upload ke Coretax. Cek kode objek pajak vs master data DJP terbaru.")
    ws["A2"].font = NOTE_FONT

    # Header row (baris 4)
    header_row = 4
    for col_idx, (label, width) in enumerate(DATA_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 30

    # Data rows (mulai baris 5)
    start = header_row + 1
    for i, r in enumerate(rows):
        row = start + i
        subtotal = float(r["subtotal"] or 0)
        rate     = float(r["pph_rate"] or 0)
        pph      = float(r["pph_amount"] or 0)
        inv_date = r["invoice_date"]

        values = [
            i + 1,
            f"{month:02d}",
            str(year),
            normalize_npwp(r["npwp"]),
            r["vendor_name"] or "",
            PPH23_OBJEK_DEFAULT,
            subtotal,
            rate,
            # PPh sebagai formula: DPP × tarif% (biar transparan & bisa dicek)
            f"=G{row}*H{row}/100",
            r["invoice_no"] or "",
            inv_date.strftime("%d/%m/%Y") if inv_date else "",
            f"PPh 23 atas {r['invoice_no']}",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            # Number format
            if col_idx in (7, 9):       # DPP, PPh
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 8:          # Tarif
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (1, 2, 3):
                cell.alignment = Alignment(horizontal="center")

    # Total row
    if rows:
        total_row = start + len(rows)
        ws.cell(row=total_row, column=6, value="TOTAL").font = Font(name="Arial", bold=True)
        c_dpp = ws.cell(row=total_row, column=7, value=f"=SUM(G{start}:G{total_row-1})")
        c_pph = ws.cell(row=total_row, column=9, value=f"=SUM(I{start}:I{total_row-1})")
        for c in (c_dpp, c_pph):
            c.font = Font(name="Arial", bold=True)
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")
            c.fill = PatternFill("solid", start_color="E8EEF7")

    ws.freeze_panes = "A5"

    # ============ SHEET 2: PANDUAN ============
    wp = wb.create_sheet("PANDUAN")
    wp.column_dimensions["A"].width = 26
    wp.column_dimensions["B"].width = 70

    wp["A1"] = "PANDUAN PENGISIAN & MAPPING KE CONVERTER DJP"
    wp["A1"].font = TITLE_FONT
    wp.merge_cells("A1:B1")

    guide = [
        ("Kolom", "Penjelasan"),
        ("Masa Pajak", "Bulan pemotongan PPh (01-12). Diisi otomatis sesuai --month."),
        ("Tahun Pajak", "Tahun pemotongan."),
        ("NPWP/NIK Pihak Dipotong", "NPWP vendor (15/16 digit, sudah dibersihkan dari titik/strip). WAJIB valid — NPWP salah = XML ditolak Coretax."),
        ("Nama Pihak Dipotong", "Nama vendor sesuai master data."),
        ("Kode Objek Pajak", f"Default {PPH23_OBJEK_DEFAULT} (jasa lain). Sesuaikan per jenis jasa — cek master data objek pajak DJP terbaru."),
        ("DPP (Rp)", "Dasar Pengenaan Pajak = subtotal invoice (sebelum PPN)."),
        ("Tarif (%)", "Tarif PPh 23 (umumnya 2%). Untuk vendor tanpa NPWP tarif jadi 2x lipat (4%) — cek manual."),
        ("PPh Dipotong (Rp)", "DPP dikali Tarif. Dihitung otomatis dengan formula Excel."),
        ("No Dokumen Referensi", "Nomor invoice sebagai referensi internal."),
        ("Tanggal Dokumen", "Tanggal invoice (dd/mm/yyyy)."),
        ("Keterangan", "Deskripsi singkat objek pemotongan."),
        ("", ""),
        ("CATATAN PENTING", ""),
        ("1. Skema XML", "Coretax 2025+ pakai import XML. Excel ini = sumber data, copy ke converter DJP lalu export XML."),
        ("2. Versi converter", "Selalu unduh converter Bupot Unifikasi TERBARU dari pajak.go.id/reformdjp/coretax (BP versi bisa berubah)."),
        ("3. NPWP tanpa identitas", "Vendor tanpa NPWP valid: tarif PPh 23 = 2x (4%). Baris seperti ini perlu dicek manual sebelum export."),
        ("4. Validasi", "Penyebab umum XML ditolak: NPWP tidak valid, kode objek pajak salah, format tanggal salah."),
        ("5. SKB / Bebas", "Vendor dengan SKB aktif TIDAK muncul di sini (pph_amount = 0, sudah ter-filter otomatis)."),
    ]
    for i, (col, desc) in enumerate(guide, start=3):
        a = wp.cell(row=i, column=1, value=col)
        b = wp.cell(row=i, column=2, value=desc)
        if col in ("Kolom", "CATATAN PENTING"):
            a.font = Font(name="Arial", bold=True, color="1F4E78")
            b.font = Font(name="Arial", bold=True, color="1F4E78")
        else:
            a.font = Font(name="Arial", bold=True, size=9)
            b.font = Font(name="Arial", size=9)
        b.alignment = Alignment(wrap_text=True, vertical="top")

    # ============ SHEET 3: REKAP ============
    wr = wb.create_sheet("REKAP")
    wr.column_dimensions["A"].width = 30
    wr.column_dimensions["B"].width = 24

    wr["A1"] = f"REKAP BUPOT PPh 23 — MASA {month:02d}/{year}"
    wr["A1"].font = TITLE_FONT
    wr.merge_cells("A1:B1")

    n = len(rows)
    data_last = start + n - 1 if n else start
    rekap = [
        ("Jumlah Bukti Potong", n),
        ("Total DPP (Rp)", f"=DATA!G{start+n}" if n else 0),
        ("Total PPh 23 (Rp)", f"=DATA!I{start+n}" if n else 0),
        ("Periode", f"{month:02d}/{year}"),
        ("Entity ID", str(entity_id) if entity_id else "SEMUA"),
        ("Digenerate", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (label, val) in enumerate(rekap, start=3):
        wr.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        c = wr.cell(row=i, column=2, value=val)
        c.font = Font(name="Arial", size=10)
        if "Rp" in label:
            c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")

    return wb

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export PPh 23 bulanan ke Excel untuk Coretax")
    parser.add_argument("--month", required=True,
                        help="Periode YYYY-MM, contoh: 2026-03")
    parser.add_argument("--entity-id", dest="entity_id", default=None,
                        help="Filter satu entity saja")
    parser.add_argument("--output", default=None,
                        help="Nama file output (default: bupot_pph23_YYYY-MM.xlsx)")
    args = parser.parse_args()

    try:
        year, month = map(int, args.month.split("-"))
    except ValueError:
        logger.error("Format --month salah. Gunakan YYYY-MM, contoh: 2026-03")
        sys.exit(1)

    output = args.output or f"bupot_pph23_{args.month}.xlsx"

    # ── Koneksi ──────────────────────────────────────────────────────────────
    try:
        engine  = create_engine(DATABASE_URL)
        Session = sessionmaker(bind=engine)
        session = Session()
        logger.info("Koneksi DB berhasil")
    except Exception as e:
        logger.error(f"Gagal koneksi DB: {e}")
        sys.exit(1)

    # ── Ambil data ───────────────────────────────────────────────────────────
    try:
        rows = get_bupot_data(session, year, month, args.entity_id)
    except Exception as e:
        logger.error(f"Query gagal: {e}")
        sys.exit(1)
    finally:
        session.close()

    logger.info(f"Ditemukan {len(rows)} bukti potong PPh 23 untuk masa {month:02d}/{year}")

    if not rows:
        print(f"\n⚠️  Tidak ada invoice PPh 23 untuk masa {month:02d}/{year}.")
        print("   Pastikan ada ap_invoice dengan pph_amount > 0 dan status approved/partial/paid.\n")
        return

    # ── Build & save ─────────────────────────────────────────────────────────
    wb = build_workbook(rows, year, month, args.entity_id)
    wb.save(output)

    total_dpp = sum(float(r["subtotal"] or 0) for r in rows)
    total_pph = sum(float(r["pph_amount"] or 0) for r in rows)

    print(f"\n{'='*58}")
    print(f"  EXPORT BUPOT PPh 23 — MASA {month:02d}/{year}")
    print(f"{'='*58}")
    print(f"  Jumlah bukti potong : {len(rows):>6}")
    print(f"  Total DPP           : Rp {total_dpp:>15,.0f}")
    print(f"  Total PPh 23        : Rp {total_pph:>15,.0f}")
    print(f"  File output         : {output}")
    print(f"{'='*58}")
    print(f"  ✅ Buka file → copy sheet DATA ke Converter XML DJP → export XML")
    print(f"{'='*58}\n")


if __name__ == "__main__":
    main()