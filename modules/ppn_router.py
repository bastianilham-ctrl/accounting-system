# modules/ppn_router.py
# PPN (Pajak Pertambahan Nilai) reconciliation + export SPT Masa PPN Excel
#
# Referensi:
#   UU PPN No.42 Tahun 2009 (sebagaimana diubah UU HPP 2021)
#   PMK 18/PMK.03/2021 : Kriteria PKP & faktur pajak
#
# Alur rekonsiliasi:
#   PPN Keluaran (output VAT) — AR invoice yang diterbitkan
#   PPN Masukan  (input VAT)  — AP invoice yang diterima
#   Kurang Bayar = PPN Keluaran - PPN Masukan > 0  → harus disetor ke DJP
#   Lebih Bayar  = PPN Masukan  - PPN Keluaran > 0 → kompensasi/restitusi

import io
from uuid import UUID
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text
from loguru import logger

from core.database import get_db

router = APIRouter(prefix="/ppn", tags=["PPN"])


# ── Endpoint: rekonsiliasi PPN masa ──────────────────────────────────────────

@router.get("/reconcile/{entity_id}")
def ppn_reconcile(
    entity_id: str,
    year:  int = Query(...),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
):
    """
    Rekonsiliasi PPN Masa: PPN Keluaran (AR) vs PPN Masukan (AP).

    Mengembalikan detail per faktur + summary kurang/lebih bayar.
    Basis data: kolom `ppn_amount` di tabel `ar_invoice` dan `ap_invoice`.
    """
    keluaran_rows = db.execute(
        text("""
            SELECT
                ai.invoice_no,
                ai.invoice_date,
                ai.customer_name,
                ai.subtotal           AS dpp,
                ai.ppn_amount,
                ai.status,
                ai.customer_npwp      AS npwp
            FROM ar_invoice ai
            WHERE ai.entity_id   = :eid
              AND ai.ppn_amount  > 0
              AND ai.status NOT IN ('cancelled')
              AND EXTRACT(YEAR  FROM ai.invoice_date) = :year
              AND EXTRACT(MONTH FROM ai.invoice_date) = :month
            ORDER BY ai.invoice_date
        """),
        {"eid": entity_id, "year": year, "month": month},
    ).fetchall()

    masukan_rows = db.execute(
        text("""
            SELECT
                ai.invoice_no,
                ai.invoice_date,
                v.vendor_name,
                ai.subtotal          AS dpp,
                ai.ppn_amount,
                ai.status,
                v.npwp
            FROM ap_invoice ai
            JOIN vendor v ON v.id = ai.vendor_id
            WHERE ai.entity_id   = :eid
              AND ai.ppn_amount  > 0
              AND ai.status NOT IN ('cancelled')
              AND EXTRACT(YEAR  FROM ai.invoice_date) = :year
              AND EXTRACT(MONTH FROM ai.invoice_date) = :month
            ORDER BY ai.invoice_date
        """),
        {"eid": entity_id, "year": year, "month": month},
    ).fetchall()

    total_keluaran = sum(float(r.ppn_amount or 0) for r in keluaran_rows)
    total_masukan  = sum(float(r.ppn_amount or 0) for r in masukan_rows)
    selisih        = total_keluaran - total_masukan

    return {
        "entity_id": entity_id,
        "masa":      f"{month:02d}/{year}",
        "ppn_keluaran": {
            "total": total_keluaran,
            "count": len(keluaran_rows),
            "detail": [_row_to_dict(r) for r in keluaran_rows],
        },
        "ppn_masukan": {
            "total": total_masukan,
            "count": len(masukan_rows),
            "detail": [_row_to_dict(r) for r in masukan_rows],
        },
        "summary": {
            "ppn_keluaran":   total_keluaran,
            "ppn_masukan":    total_masukan,
            "selisih":        selisih,
            "status":         "kurang_bayar" if selisih > 0 else "lebih_bayar" if selisih < 0 else "nihil",
            "keterangan": (
                f"Kurang bayar Rp {abs(selisih):,.0f} — setor via SSP/MPN ke DJP"
                if selisih > 0 else
                f"Lebih bayar Rp {abs(selisih):,.0f} — dapat dikompensasi masa berikutnya"
                if selisih < 0 else
                "PPN nihil"
            ),
        },
    }


# ── Endpoint: YTD rekap per bulan ─────────────────────────────────────────────

@router.get("/ytd/{entity_id}")
def ppn_ytd(
    entity_id: str,
    year: int = Query(default=date.today().year),
    db: Session = Depends(get_db),
):
    """Ringkasan PPN per bulan sepanjang tahun."""
    rows = db.execute(
        text("""
            SELECT
                bulan,
                SUM(ppn_keluaran) AS ppn_keluaran,
                SUM(ppn_masukan)  AS ppn_masukan
            FROM (
                SELECT
                    EXTRACT(MONTH FROM invoice_date)::int AS bulan,
                    ppn_amount                            AS ppn_keluaran,
                    0                                     AS ppn_masukan
                FROM ar_invoice
                WHERE entity_id = :eid
                  AND EXTRACT(YEAR FROM invoice_date) = :year
                  AND ppn_amount > 0
                  AND status NOT IN ('cancelled')

                UNION ALL

                SELECT
                    EXTRACT(MONTH FROM invoice_date)::int AS bulan,
                    0                                     AS ppn_keluaran,
                    ppn_amount                            AS ppn_masukan
                FROM ap_invoice
                WHERE entity_id = :eid
                  AND EXTRACT(YEAR FROM invoice_date) = :year
                  AND ppn_amount > 0
                  AND status NOT IN ('cancelled')
            ) t
            GROUP BY bulan
            ORDER BY bulan
        """),
        {"eid": entity_id, "year": year},
    ).fetchall()

    months = {r.bulan: {"keluaran": float(r.ppn_keluaran), "masukan": float(r.ppn_masukan)} for r in rows}
    result = []
    for m in range(1, 13):
        d = months.get(m, {"keluaran": 0, "masukan": 0})
        selisih = d["keluaran"] - d["masukan"]
        result.append({
            "bulan":        m,
            "nama_bulan":   _month_name(m),
            "ppn_keluaran": d["keluaran"],
            "ppn_masukan":  d["masukan"],
            "selisih":      selisih,
            "status":       "kurang_bayar" if selisih > 0 else "lebih_bayar" if selisih < 0 else "nihil",
        })

    total_k = sum(r["ppn_keluaran"] for r in result)
    total_m = sum(r["ppn_masukan"]  for r in result)
    return {
        "entity_id": entity_id,
        "year":      year,
        "per_bulan": result,
        "ytd_summary": {
            "total_ppn_keluaran": total_k,
            "total_ppn_masukan":  total_m,
            "net_ppn":            total_k - total_m,
        },
    }


# ── Endpoint: export SPT Masa PPN Excel ──────────────────────────────────────

@router.get("/export-spt/{entity_id}")
def export_spt_ppn(
    entity_id: str,
    year:  int = Query(...),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
):
    """
    Export SPT Masa PPN ke Excel — format referensi Formulir 1111.

    Sheet 1: REKAP   — ringkasan kurang/lebih bayar
    Sheet 2: KELUARAN — detail faktur pajak keluaran (PK)
    Sheet 3: MASUKAN  — detail faktur pajak masukan (PM)
    Sheet 4: PANDUAN  — petunjuk pengisian
    """
    if not (1 <= month <= 12):
        raise HTTPException(400, "month harus antara 1-12")

    keluaran_rows = db.execute(
        text("""
            SELECT
                ai.invoice_no,
                ai.invoice_date,
                ai.customer_name,
                ai.customer_npwp       AS npwp_lawan,
                ai.subtotal            AS dpp,
                ai.ppn_amount,
                ai.status,
                'Keluaran'             AS jenis
            FROM ar_invoice ai
            WHERE ai.entity_id   = :eid
              AND ai.ppn_amount  > 0
              AND ai.status NOT IN ('cancelled')
              AND EXTRACT(YEAR  FROM ai.invoice_date) = :year
              AND EXTRACT(MONTH FROM ai.invoice_date) = :month
            ORDER BY ai.invoice_date
        """),
        {"eid": entity_id, "year": year, "month": month},
    ).fetchall()

    masukan_rows = db.execute(
        text("""
            SELECT
                ai.invoice_no,
                ai.invoice_date,
                v.vendor_name          AS customer_name,
                v.npwp                 AS npwp_lawan,
                ai.subtotal            AS dpp,
                ai.ppn_amount,
                ai.status,
                'Masukan'              AS jenis
            FROM ap_invoice ai
            JOIN vendor v ON v.id = ai.vendor_id
            WHERE ai.entity_id   = :eid
              AND ai.ppn_amount  > 0
              AND ai.status NOT IN ('cancelled')
              AND EXTRACT(YEAR  FROM ai.invoice_date) = :year
              AND EXTRACT(MONTH FROM ai.invoice_date) = :month
            ORDER BY ai.invoice_date
        """),
        {"eid": entity_id, "year": year, "month": month},
    ).fetchall()

    total_keluaran = sum(float(r.ppn_amount or 0) for r in keluaran_rows)
    total_masukan  = sum(float(r.ppn_amount or 0) for r in masukan_rows)

    wb = _build_spt_excel(
        keluaran_rows, masukan_rows,
        total_keluaran, total_masukan,
        year, month, entity_id,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"spt_masa_ppn_{year}-{month:02d}.xlsx"
    logger.info(
        f"SPT Masa PPN export: masa {month:02d}/{year} | "
        f"keluaran {len(keluaran_rows)} | masukan {len(masukan_rows)}"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Helpers ───────────────────────────────────────────────────────────────────

def _row_to_dict(r) -> dict:
    d = {}
    for k in r._fields:
        v = getattr(r, k)
        d[k] = str(v) if isinstance(v, (date, datetime, UUID)) else v
    return d


def _month_name(m: int) -> str:
    names = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
    return names[m - 1]


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_spt_excel(keluaran_rows, masukan_rows, total_k, total_m, year, month, entity_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY  = PatternFill("solid", start_color="1F4E78")
    GREEN = PatternFill("solid", start_color="1E5631")
    HDR_F = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TTL_F = Font(name="Arial", bold=True, size=13, color="1F4E78")
    LBL_F = Font(name="Arial", bold=True, size=10)
    CEL_F = Font(name="Arial", size=10)
    THIN  = Side(style="thin", color="CCCCCC")
    BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    selisih = total_k - total_m
    wb = Workbook()

    # ── Sheet REKAP ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "REKAP"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24

    ws.merge_cells("A1:B1")
    ws["A1"] = f"SPT MASA PPN — MASA {month:02d}/{year}"
    ws["A1"].font = TTL_F

    rekap_data = [
        ("KOMPONEN PPN",               "JUMLAH (Rp)"),
        ("PPN Keluaran (output VAT)",   total_k),
        ("PPN Masukan (input VAT)",     total_m),
        ("",                            ""),
        ("PPN Kurang/Lebih Bayar",      selisih),
        ("Status",
            "KURANG BAYAR — Setor ke DJP" if selisih > 0
            else "LEBIH BAYAR — Kompensasi/Restitusi" if selisih < 0
            else "NIHIL"),
        ("",                            ""),
        ("Masa Pajak",                  f"{month:02d}/{year}"),
        ("Jumlah FK Keluaran",          len(keluaran_rows)),
        ("Jumlah FK Masukan",           len(masukan_rows)),
        ("Entity ID",                   entity_id),
        ("Digenerate",                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for i, (lbl, val) in enumerate(rekap_data, start=3):
        ca = ws.cell(row=i, column=1, value=lbl)
        cb = ws.cell(row=i, column=2, value=val)
        bold = lbl in ("KOMPONEN PPN", "PPN Kurang/Lebih Bayar", "Status")
        ca.font = Font(name="Arial", bold=bold, size=10)
        cb.font = Font(name="Arial", bold=bold, size=10)
        if isinstance(val, (int, float)) and "Status" not in lbl and "Jumlah FK" not in lbl and "Entity" not in lbl:
            cb.number_format = "#,##0"
            cb.alignment = Alignment(horizontal="right")
        if lbl == "PPN Kurang/Lebih Bayar":
            color = "FFC7CE" if selisih > 0 else "C7E6C7" if selisih < 0 else "FFFFFF"
            cb.fill = PatternFill("solid", start_color=color)

    # ── Sheet KELUARAN ────────────────────────────────────────────────────────
    wk = wb.create_sheet("KELUARAN")
    _write_faktur_sheet(
        ws=wk,
        rows=keluaran_rows,
        title=f"FAKTUR PAJAK KELUARAN (PK) — MASA {month:02d}/{year}",
        counterparty_label="Nama Pembeli",
        fill=NAVY,
        hdr_font=HDR_F,
        cell_font=CEL_F,
        border=BDR,
    )

    # ── Sheet MASUKAN ─────────────────────────────────────────────────────────
    wm = wb.create_sheet("MASUKAN")
    _write_faktur_sheet(
        ws=wm,
        rows=masukan_rows,
        title=f"FAKTUR PAJAK MASUKAN (PM) — MASA {month:02d}/{year}",
        counterparty_label="Nama Penjual/Pemasok",
        fill=GREEN,
        hdr_font=HDR_F,
        cell_font=CEL_F,
        border=BDR,
    )

    # ── Sheet PANDUAN ─────────────────────────────────────────────────────────
    wp = wb.create_sheet("PANDUAN")
    wp.column_dimensions["A"].width = 30
    wp.column_dimensions["B"].width = 70
    wp["A1"] = "PANDUAN SPT MASA PPN"
    wp["A1"].font = TTL_F
    wp.merge_cells("A1:B1")

    panduan = [
        ("Kolom / Term",            "Penjelasan"),
        ("PPN Keluaran (PK)",       "PPN yang dipungut dari pelanggan (Pembeli). Berasal dari AR invoice. 11% × DPP."),
        ("PPN Masukan (PM)",        "PPN yang dibayar ke pemasok (AP invoice). Dapat dikreditkan."),
        ("DPP",                     "Dasar Pengenaan Pajak = nilai transaksi sebelum PPN."),
        ("FK",                      "Faktur Pajak — dokumen yang harus diterbitkan saat transaksi."),
        ("NPWP Lawan Transaksi",    "NPWP pembeli (PK) atau penjual (PM). Isi 15 digit tanpa strip."),
        ("",                        ""),
        ("CARA LAPOR",              ""),
        ("1. Login",                "Akses DJP Online (djponline.pajak.go.id) → e-Filing → SPT Masa PPN"),
        ("2. Impor data",           "Gunakan Formulir 1111 atau e-Faktur untuk upload data detail"),
        ("3. Verifikasi",           "Pastikan total PK dan PM di sistem sama dengan e-Faktur"),
        ("4. Submit",               "Tanda tangani elektronik lalu submit sebelum akhir bulan berikutnya"),
        ("",                        ""),
        ("CATATAN",                 ""),
        ("Tarif PPN",               "11% berlaku sejak 1 April 2022 (PMK 60/PMK.03/2022 / UU HPP 2021)"),
        ("NSFP",                    "Nomor Seri Faktur Pajak harus diminta ke KPP sebelum penerbitan FK"),
        ("Tanggal jatuh tempo",     "SPT Masa PPN paling lambat dilaporkan akhir bulan berikutnya"),
    ]
    for i, (a, b) in enumerate(panduan, start=3):
        ca = wp.cell(row=i, column=1, value=a)
        cb = wp.cell(row=i, column=2, value=b)
        bold = a in ("Kolom / Term", "CARA LAPOR", "CATATAN")
        ca.font = Font(name="Arial", bold=bold, size=9)
        cb.font = Font(name="Arial", size=9, bold=bold)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    return wb


def _write_faktur_sheet(ws, rows, title, counterparty_label, fill, hdr_font, cell_font, border):
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    TITLE_FONT = Font(name="Arial", bold=True, size=12, color="1F4E78")
    NOTE_FONT  = Font(name="Arial", italic=True, size=9, color="808080")

    COLS = [
        ("No",                    5),
        ("No Faktur / Invoice",  22),
        ("Tanggal Faktur",       15),
        (counterparty_label,     32),
        ("NPWP Lawan Transaksi", 22),
        ("DPP (Rp)",             18),
        ("PPN (Rp)",             18),
        ("Status",               14),
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")
    ws["A2"] = "Data ini adalah rangkuman dari sistem. Pastikan sudah diinput ke e-Faktur sebelum lapor SPT."
    ws["A2"].font = NOTE_FONT

    HR = 4
    for ci, (label, width) in enumerate(COLS, start=1):
        c = ws.cell(row=HR, column=ci, value=label)
        c.font      = hdr_font
        c.fill      = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HR].height = 32

    START = HR + 1
    for i, r in enumerate(rows):
        row   = START + i
        npwp  = "".join(c for c in str(r.npwp_lawan or "") if c.isdigit())
        inv_dt = r.invoice_date

        vals = [
            i + 1,
            r.invoice_no or "",
            inv_dt.strftime("%d/%m/%Y") if inv_dt else "",
            r.customer_name or "",
            npwp,
            float(r.dpp or 0),
            float(r.ppn_amount or 0),
            r.status or "",
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font   = cell_font
            c.border = border
            if ci in (6, 7):
                c.number_format = "#,##0"
                c.alignment     = Alignment(horizontal="right")
            elif ci in (1, 3, 8):
                c.alignment     = Alignment(horizontal="center")

    # Total row
    if rows:
        tr  = START + len(rows)
        lbl = get_column_letter(len(COLS) - 2)
        ws.cell(row=tr, column=5, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
        for ci, col in [(6, "F"), (7, "G")]:
            c = ws.cell(row=tr, column=ci, value=f"=SUM({col}{START}:{col}{tr-1})")
            c.font          = Font(name="Arial", bold=True)
            c.number_format = "#,##0"
            c.alignment     = Alignment(horizontal="right")
        from openpyxl.styles import PatternFill
        for ci in range(1, len(COLS) + 1):
            ws.cell(row=tr, column=ci).fill = PatternFill("solid", start_color="E8EEF7")

    ws.freeze_panes = "A5"
