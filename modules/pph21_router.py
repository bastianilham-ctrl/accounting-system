# modules/pph21_router.py
# PPh 21 endpoints: kalkulasi, ringkasan YTD per payee, export Bupot 21 Excel

import io
from uuid import UUID
from decimal import Decimal
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text
from loguru import logger

from core.database import get_db
from modules.pph21_engine import PPh21Engine, calculate_pph21, is_tenaga_ahli, TENAGA_AHLI_KEYWORDS

router = APIRouter(prefix="/pph21", tags=["PPh 21"])


# ── Request models ─────────────────────────────────────────────────────────────

class PPh21CalcRequest(BaseModel):
    gross_amount:       Decimal
    has_npwp:           bool    = True
    is_tenaga_ahli:     bool    = True
    ytd_gross_before:   Decimal = Decimal("0")  # akumulasi YTD sebelum pembayaran ini


class PPh21InvoiceCalcRequest(BaseModel):
    vendor_id:    UUID
    gross_amount: Decimal
    description:  str
    invoice_date: date


# ── Endpoints ──────────────────────────────────────────────────────────────────

@router.post("/calculate")
def calculate_pph21_standalone(req: PPh21CalcRequest):
    """
    Hitung PPh 21 tanpa lookup DB — untuk kalkulasi ad-hoc / simulasi.

    **Cara hitung:**
    - Tenaga ahli: PKP = 50% × bruto, lalu tarif progresif Pasal 17
    - Bukan tenaga ahli: PKP = bruto
    - Tanpa NPWP: PPh × 1.20 (Pasal 21 ayat 5a UU PPh)

    **Tarif progresif (UU HPP 2021 Pasal 17):**
    | PKP               | Rate |
    |-------------------|------|
    | s/d Rp 60 juta    | 5%   |
    | Rp 60–250 juta    | 15%  |
    | Rp 250–500 juta   | 25%  |
    | Rp 500 juta–5 M   | 30%  |
    | > Rp 5 milyar     | 35%  |
    """
    return calculate_pph21(
        gross_amount      = req.gross_amount,
        has_npwp          = req.has_npwp,
        is_tenaga_ahli    = req.is_tenaga_ahli,
        ytd_gross_before  = req.ytd_gross_before,
    )


@router.post("/calculate-for-vendor")
def calculate_pph21_for_vendor(req: PPh21InvoiceCalcRequest, db: Session = Depends(get_db)):
    """
    Hitung PPh 21 untuk vendor individu (Perorangan) dari database.
    Otomatis:
    - Cek NPWP dari master vendor
    - Deteksi tenaga ahli dari deskripsi
    - Hitung YTD income dari invoice sebelumnya tahun ini
    """
    engine = PPh21Engine(db)
    result = engine.calculate_for_invoice(
        vendor_id    = req.vendor_id,
        gross_amount = req.gross_amount,
        description  = req.description,
        invoice_date = req.invoice_date,
    )
    if "error" in result:
        raise HTTPException(404, result["error"])
    return result


@router.get("/summary/{entity_id}")
def pph21_ytd_summary(
    entity_id: str,
    year: int = Query(default=date.today().year),
    db: Session = Depends(get_db),
):
    """
    Ringkasan PPh 21 YTD per payee individu.
    Menampilkan total bruto, total PPh 21 dipotong, dan jumlah invoice per orang.
    """
    engine = PPh21Engine(db)
    rows   = engine.get_ytd_summary(UUID(entity_id), year)

    total_bruto = sum(float(r["total_bruto"] or 0) for r in rows)
    total_pph   = sum(float(r["total_pph21"] or 0) for r in rows)

    return {
        "entity_id":  entity_id,
        "year":       year,
        "payees":     rows,
        "summary": {
            "total_payees":    len(rows),
            "total_bruto":     total_bruto,
            "total_pph21":     total_pph,
            "effective_rate":  round(total_pph / total_bruto * 100, 2) if total_bruto > 0 else 0,
        },
    }


@router.get("/bupot/{entity_id}")
def export_bupot21(
    entity_id: str,
    year:  int = Query(..., description="Tahun pajak"),
    month: int = Query(..., description="Masa pajak (1-12)"),
    db:    Session = Depends(get_db),
):
    """
    Export Bukti Potong PPh 21 bulanan ke Excel.
    Format siap copy ke Converter XML Bupot Unifikasi DJP (Coretax).

    Hanya transaksi dengan pph_type LIKE 'PPh21%', pph_amount > 0,
    dan status approved/partial/paid yang masuk.
    """
    if not (1 <= month <= 12):
        raise HTTPException(400, "month harus antara 1-12")

    rows = db.execute(
        text("""
            SELECT
                ai.invoice_no,
                ai.invoice_date,
                ai.subtotal       AS bruto,
                ai.pph_amount     AS pph21,
                ai.pph_rate,
                ai.pph_type,
                v.vendor_name,
                v.npwp,
                v.vendor_code
            FROM ap_invoice ai
            JOIN vendor v ON v.id = ai.vendor_id
            WHERE ai.entity_id  = :eid
              AND ai.pph_type LIKE 'PPh21%'
              AND ai.pph_amount > 0
              AND ai.status IN ('approved','partial','paid')
              AND EXTRACT(YEAR  FROM ai.invoice_date) = :year
              AND EXTRACT(MONTH FROM ai.invoice_date) = :month
            ORDER BY ai.invoice_date, v.vendor_name
        """),
        {"eid": entity_id, "year": year, "month": month}
    ).fetchall()

    wb = _build_bupot21_excel(rows, year, month, entity_id)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bupot_pph21_{year}-{month:02d}.xlsx"
    logger.info(f"Bupot PPh 21 export: {len(rows)} baris | masa {month:02d}/{year}")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reference/tenaga-ahli-keywords")
def tenaga_ahli_keywords():
    """Daftar keyword yang digunakan untuk mendeteksi tenaga ahli."""
    return {"keywords": TENAGA_AHLI_KEYWORDS}


# ── Excel builder ──────────────────────────────────────────────────────────────

def _build_bupot21_excel(rows, year: int, month: int, entity_id: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", start_color="1F4E78")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE_FONT  = Font(name="Arial", bold=True, size=13, color="1F4E78")
    NOTE_FONT   = Font(name="Arial", italic=True, size=9, color="808080")
    CELL_FONT   = Font(name="Arial", size=10)
    THIN        = Side(style="thin", color="D0D0D0")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Kode objek pajak PPh 21 default (honorarium tenaga ahli)
    # 21-100-99: Penghasilan lain (bukan pegawai)
    # Sesuaikan dengan master data DJP terbaru saat upload ke Coretax
    OBJEK_DEFAULT = "21-100-99"

    DATA_COLS = [
        ("No",                       6),
        ("Masa Pajak",              12),
        ("Tahun Pajak",             12),
        ("NIK/NPWP Pihak Dipotong", 24),
        ("Nama Pihak Dipotong",     30),
        ("Kode Objek Pajak",        16),
        ("DPP (Bruto, Rp)",         18),
        ("PKP (50% DPP, Rp)",       18),
        ("Tarif (%)",               10),
        ("PPh 21 Dipotong (Rp)",    20),
        ("No Dokumen",              20),
        ("Tanggal Dokumen",         16),
        ("Keterangan",              28),
    ]

    wb = Workbook()

    # ── Sheet DATA ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "DATA"

    ws.merge_cells("A1:M1")
    ws["A1"] = f"DAFTAR BUKTI POTONG PPh PASAL 21 — MASA {month:02d}/{year}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        "Copy baris data (mulai baris 5) ke Converter XML Bupot Unifikasi DJP, "
        "lalu export XML untuk upload ke Coretax. Cek kode objek pajak vs master DJP terbaru."
    )
    ws["A2"].font = NOTE_FONT

    # Header row 4
    HR = 4
    for ci, (label, width) in enumerate(DATA_COLS, start=1):
        c = ws.cell(row=HR, column=ci, value=label)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HR].height = 36

    # Data rows
    START = HR + 1
    for i, r in enumerate(rows):
        row    = START + i
        bruto  = float(r.bruto or 0)
        pkp    = round(bruto * 0.5, 0)      # tenaga ahli default 50%
        pph    = float(r.pph21 or 0)
        rate   = float(r.pph_rate or 0)
        npwp   = "".join(c for c in str(r.npwp or "") if c.isdigit())
        inv_dt = r.invoice_date

        vals = [
            i + 1,
            f"{month:02d}",
            str(year),
            npwp,
            r.vendor_name or "",
            OBJEK_DEFAULT,
            bruto,
            pkp,
            rate,
            f"=J{row}",          # formula: PPh dari kolom J (sesuaikan jika beda)
            r.invoice_no or "",
            inv_dt.strftime("%d/%m/%Y") if inv_dt else "",
            f"PPh 21 atas {r.invoice_no}",
        ]
        for ci, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font   = CELL_FONT
            c.border = BORDER
            if ci in (7, 8, 10):
                c.number_format = "#,##0"
                c.alignment     = Alignment(horizontal="right")
            elif ci == 9:
                c.number_format = "0.00"
                c.alignment     = Alignment(horizontal="center")
            elif ci in (1, 2, 3):
                c.alignment     = Alignment(horizontal="center")

    # Total row
    if rows:
        tr = START + len(rows)
        ws.cell(row=tr, column=6, value="TOTAL").font = Font(name="Arial", bold=True)
        for ci, col in [(7, "G"), (8, "H"), (10, "J")]:
            c = ws.cell(row=tr, column=ci,
                        value=f"=SUM({col}{START}:{col}{tr-1})")
            c.font         = Font(name="Arial", bold=True)
            c.number_format = "#,##0"
            c.alignment    = Alignment(horizontal="right")
            c.fill         = PatternFill("solid", start_color="E8EEF7")

    ws.freeze_panes = "A5"

    # ── Sheet PANDUAN ─────────────────────────────────────────────────────────
    wp = wb.create_sheet("PANDUAN")
    wp.column_dimensions["A"].width = 28
    wp.column_dimensions["B"].width = 72

    wp["A1"] = "PANDUAN PPh 21 BUKAN PEGAWAI"
    wp["A1"].font = TITLE_FONT
    wp.merge_cells("A1:B1")

    guide = [
        ("Kolom",                   "Keterangan"),
        ("NIK/NPWP Pihak Dipotong", "NPWP (15 digit) atau NIK (16 digit) payee. WAJIB valid — salah = XML ditolak."),
        ("Kode Objek Pajak",        f"Default {OBJEK_DEFAULT}. Sesuaikan: 21-100-01 (honorarium PNS), 21-100-05 (hadiah), dll."),
        ("DPP (Bruto)",             "Penghasilan bruto sebelum dipotong PPh."),
        ("PKP (50% DPP)",           "Dasar pengenaan untuk tenaga ahli = 50% × bruto (PMK 168/2023 Pasal 14)."),
        ("Tarif",                   "Tarif progresif Pasal 17: 5%/15%/25%/30%/35%. Kolom ini = tarif efektif actual."),
        ("PPh 21 Dipotong",         "Nilai PPh 21 yang dipotong. Dihitung saat posting invoice."),
        ("",                        ""),
        ("CATATAN PENTING",         ""),
        ("1. Tanpa NPWP",           "Tarif × 1.20 (Pasal 21 ayat 5a UU PPh). Isi NIK 16 digit jika ada."),
        ("2. Tarif progresif",      "5% s/d Rp 60 jt | 15% Rp 60-250 jt | 25% Rp 250-500 jt | 30% Rp 500 jt-5 M | 35% >5 M"),
        ("3. YTD cumulative",       "Sistem memperhitungkan akumulasi penghasilan ke payee yang sama dalam satu tahun."),
        ("4. Tenaga ahli",          "Konsultan, notaris, pengacara, dokter, akuntan, arsitek, penilai, aktuaris."),
        ("5. Upload Coretax",       "Unduh converter Bupot Unifikasi TERBARU di pajak.go.id/reformdjp/coretax."),
    ]
    for i, (a, b) in enumerate(guide, start=3):
        ca = wp.cell(row=i, column=1, value=a)
        cb = wp.cell(row=i, column=2, value=b)
        bold = a in ("Kolom", "CATATAN PENTING")
        ca.font = Font(name="Arial", bold=bold, size=9)
        cb.font = Font(name="Arial", bold=bold, size=9)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Sheet REKAP ───────────────────────────────────────────────────────────
    wr = wb.create_sheet("REKAP")
    wr.column_dimensions["A"].width = 32
    wr.column_dimensions["B"].width = 24

    wr["A1"] = f"REKAP BUPOT PPh 21 — MASA {month:02d}/{year}"
    wr["A1"].font = TITLE_FONT
    wr.merge_cells("A1:B1")

    n      = len(rows)
    s      = START
    rekap  = [
        ("Jumlah Bukti Potong",    n),
        ("Total DPP Bruto (Rp)",   f"=DATA!G{s+n}" if n else 0),
        ("Total PPh 21 (Rp)",      f"=DATA!J{s+n}" if n else 0),
        ("Masa Pajak",             f"{month:02d}/{year}"),
        ("Entity ID",              entity_id),
        ("Digenerate",             datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (lbl, val) in enumerate(rekap, start=3):
        wr.cell(row=i, column=1, value=lbl).font = Font(name="Arial", bold=True, size=10)
        c = wr.cell(row=i, column=2, value=val)
        c.font = Font(name="Arial", size=10)
        if "Rp" in lbl:
            c.number_format = "#,##0"
            c.alignment     = Alignment(horizontal="right")

    return wb
