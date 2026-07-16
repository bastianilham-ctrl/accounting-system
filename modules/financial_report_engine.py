"""
Financial Report Engine
========================
Lima laporan keuangan formal:

  1. Trial Balance      — saldo semua akun per periode
  2. Balance Sheet      — neraca per tanggal tertentu
  3. Profit & Loss      — laba rugi untuk rentang tanggal
  4. General Ledger     — buku besar detail per akun
  5. Cash Flow          — arus kas metode tidak langsung

Prinsip:
  - Semua laporan mengacu ke gl_line (debit_idr / credit_idr)
  - Balance sheet + P&L harus balance (Aktiva = Pasiva)
  - Net Income di P&L masuk ke Equity sebagai "Laba Periode Berjalan"
  - Cash Flow dimulai dari Net Income lalu penyesuaian working capital

Klasifikasi akun berdasarkan account_type di chart_of_accounts:
  asset     → Aktiva
  liability → Kewajiban
  equity    → Ekuitas
  revenue   → Pendapatan
  cogs      → HPP
  expense   → Beban Operasional
  other_income   → Pendapatan Lain-lain   (account_code prefix 7-)
  other_expense  → Beban Lain-lain        (account_code prefix 8-)
  tax_expense    → Beban Pajak            (account_code prefix 9-)

Export: PDF (WeasyPrint) dan Excel (openpyxl)
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional
from sqlalchemy import text
from sqlalchemy.orm import Session


# ── Helpers ───────────────────────────────────────────────────────────────────

def _d(val) -> Decimal:
    return Decimal(str(val)) if val is not None else Decimal("0")


def _fmt(val) -> str:
    """Format angka Rupiah: 1234567 → '1.234.567'"""
    try:
        n = int(round(float(val)))
        if n < 0:
            return f"({abs(n):,})".replace(",", ".")
        return f"{n:,}".replace(",", ".")
    except Exception:
        return str(val)


def _last_day_of_month(year: int, month: int) -> date:
    import calendar
    return date(year, month, calendar.monthrange(year, month)[1])


# ── 1. TRIAL BALANCE ─────────────────────────────────────────────────────────

class FinancialReportEngine:

    @staticmethod
    def get_trial_balance(
        db: Session,
        entity_id: str,
        fiscal_year: int,
        fiscal_month: int,
        include_zero_balance: bool = False,
        compare_year: Optional[int] = None,
        compare_month: Optional[int] = None,
    ) -> dict:
        """
        Trial Balance per periode fiskal.
        Menampilkan: Opening Balance, Mutasi Debit/Kredit, Closing Balance.
        Comparative: bandingkan dengan periode lain (opsional).
        """
        as_of = _last_day_of_month(fiscal_year, fiscal_month)
        period_start = date(fiscal_year, fiscal_month, 1)

        def _fetch(year: int, month: int) -> list:
            period_s = date(year, month, 1)
            period_e = _last_day_of_month(year, month)
            prior_e  = period_s - timedelta(days=1)

            rows = db.execute(text("""
                SELECT
                    coa.account_code,
                    coa.account_name,
                    coa.account_type,
                    coa.normal_balance,

                    -- Opening: semua transaksi SEBELUM periode ini
                    COALESCE(SUM(gl.debit_idr)  FILTER (WHERE gj.journal_date < :ps), 0) AS open_dr,
                    COALESCE(SUM(gl.credit_idr) FILTER (WHERE gj.journal_date < :ps), 0) AS open_cr,

                    -- Period: transaksi DALAM periode ini
                    COALESCE(SUM(gl.debit_idr)  FILTER (WHERE gj.journal_date BETWEEN :ps AND :pe), 0) AS per_dr,
                    COALESCE(SUM(gl.credit_idr) FILTER (WHERE gj.journal_date BETWEEN :ps AND :pe), 0) AS per_cr

                FROM chart_of_accounts coa
                LEFT JOIN gl_line gl  ON gl.account_id = coa.id
                LEFT JOIN gl_journal gj ON gj.id = gl.journal_id
                                       AND gj.status = 'posted'
                                       AND gj.entity_id = :eid
                                       AND gj.journal_date <= :pe
                WHERE coa.entity_id = :eid
                  AND coa.is_header = FALSE
                GROUP BY coa.account_code, coa.account_name, coa.account_type, coa.normal_balance
                ORDER BY coa.account_code
            """), {"eid": entity_id, "ps": period_s, "pe": period_e}).fetchall()
            return rows

        rows = _fetch(fiscal_year, fiscal_month)

        accounts = []
        total_open_dr = total_open_cr = Decimal("0")
        total_per_dr = total_per_cr = Decimal("0")
        total_close_dr = total_close_cr = Decimal("0")

        for r in rows:
            open_dr  = _d(r.open_dr)
            open_cr  = _d(r.open_cr)
            per_dr   = _d(r.per_dr)
            per_cr   = _d(r.per_cr)
            close_dr = open_dr + per_dr
            close_cr = open_cr + per_cr

            # Saldo neto: Dr - Cr atau Cr - Dr tergantung normal balance
            normal = (r.normal_balance or "debit").lower()
            if normal == "debit":
                closing_balance = close_dr - close_cr
                close_dr_show = max(closing_balance, Decimal("0"))
                close_cr_show = max(-closing_balance, Decimal("0"))
            else:
                closing_balance = close_cr - close_dr
                close_cr_show = max(closing_balance, Decimal("0"))
                close_dr_show = max(-closing_balance, Decimal("0"))

            if not include_zero_balance and close_dr_show == 0 and close_cr_show == 0:
                continue

            accounts.append({
                "account_code":   r.account_code,
                "account_name":   r.account_name,
                "account_type":   r.account_type,
                "normal_balance": normal,
                "opening_debit":  float(open_dr),
                "opening_credit": float(open_cr),
                "period_debit":   float(per_dr),
                "period_credit":  float(per_cr),
                "closing_debit":  float(close_dr_show),
                "closing_credit": float(close_cr_show),
            })

            total_open_dr  += open_dr
            total_open_cr  += open_cr
            total_per_dr   += per_dr
            total_per_cr   += per_cr
            total_close_dr += close_dr_show
            total_close_cr += close_cr_show

        return {
            "entity_id":   entity_id,
            "fiscal_year": fiscal_year,
            "fiscal_month": fiscal_month,
            "as_of_date":  str(as_of),
            "accounts":    accounts,
            "totals": {
                "opening_debit":  float(total_open_dr),
                "opening_credit": float(total_open_cr),
                "period_debit":   float(total_per_dr),
                "period_credit":  float(total_per_cr),
                "closing_debit":  float(total_close_dr),
                "closing_credit": float(total_close_cr),
                "is_balanced":    abs(total_close_dr - total_close_cr) < Decimal("2"),
                "difference":     float(total_close_dr - total_close_cr),
            },
        }

    # ── 2. BALANCE SHEET (NERACA) ─────────────────────────────────────────────

    @staticmethod
    def get_balance_sheet(
        db: Session,
        entity_id: str,
        as_of_date: date,
        compare_date: Optional[date] = None,
    ) -> dict:
        """
        Neraca per tanggal tertentu.
        Net Income dari P&L dimasukkan ke Ekuitas sebagai 'Laba Periode Berjalan'.
        Mendukung kolom pembanding (compare_date).
        """
        def _fetch_balances(upto: date) -> dict:
            """Hitung saldo neto setiap akun sampai tanggal upto."""
            rows = db.execute(text("""
                SELECT
                    coa.account_code,
                    coa.account_name,
                    coa.account_type,
                    coa.normal_balance,
                    coa.account_group,
                    COALESCE(SUM(gl.debit_idr), 0)  AS total_dr,
                    COALESCE(SUM(gl.credit_idr), 0) AS total_cr
                FROM chart_of_accounts coa
                LEFT JOIN gl_line gl  ON gl.account_id = coa.id
                LEFT JOIN gl_journal gj ON gj.id = gl.journal_id
                                       AND gj.status = 'posted'
                                       AND gj.entity_id = :eid
                                       AND gj.journal_date <= :upto
                WHERE coa.entity_id = :eid
                  AND coa.is_header = FALSE
                GROUP BY coa.account_code, coa.account_name, coa.account_type,
                         coa.normal_balance, coa.account_group
                ORDER BY coa.account_code
            """), {"eid": entity_id, "upto": upto}).fetchall()

            result = {}
            for r in rows:
                dr = _d(r.total_dr)
                cr = _d(r.total_cr)
                normal = (r.normal_balance or "debit").lower()
                balance = (dr - cr) if normal == "debit" else (cr - dr)
                result[r.account_code] = {
                    "account_code":  r.account_code,
                    "account_name":  r.account_name,
                    "account_type":  r.account_type,
                    "normal_balance": normal,
                    "account_group": r.account_group,
                    "balance":       balance,
                }
            return result

        current = _fetch_balances(as_of_date)
        compare = _fetch_balances(compare_date) if compare_date else {}

        # Pisahkan net income dari P&L (revenue - cogs - expense - other_expense + other_income - tax)
        pl_types = {"revenue", "cogs", "expense", "other_income", "other_expense", "tax_expense"}

        def _net_income(balances: dict) -> Decimal:
            net = Decimal("0")
            for acct in balances.values():
                t = acct["account_type"]
                b = acct["balance"]
                if t == "revenue":        net += b
                elif t == "cogs":         net -= b
                elif t == "expense":      net -= b
                elif t == "other_income": net += b
                elif t == "other_expense": net -= b
                elif t == "tax_expense":  net -= b
            return net

        curr_net_income = _net_income(current)
        comp_net_income = _net_income(compare) if compare else Decimal("0")

        def _build_sections(balances: dict, net_income: Decimal) -> dict:
            """
            Bangun struktur Aktiva / Pasiva dari balances.
            """
            # Klasifikasi ke grup
            ASSET_GROUPS = {
                "1-1": "Kas dan Setara Kas",
                "1-2": "Piutang Usaha",
                "1-3": "Persediaan",
                "1-4": "Aset Lancar Lainnya",
                "1-5": "Aset Tetap",
                "1-6": "Akumulasi Penyusutan",
                "1-7": "Aset Tidak Berwujud",
                "1-8": "Investasi Jangka Panjang",
                "1-9": "Piutang Antar Perusahaan",
            }
            CURRENT_ASSET_PREFIXES = {"1-1", "1-2", "1-3", "1-4", "1-9"}
            FIXED_ASSET_PREFIXES   = {"1-5", "1-6", "1-7", "1-8"}

            LIA_GROUPS = {
                "2-1": "Hutang Usaha",
                "2-2": "Hutang Akrual dan Gaji",
                "2-3": "Hutang Bank Jangka Pendek",
                "2-4": "Kewajiban Jangka Panjang",
                "2-9": "Hutang Antar Perusahaan",
            }
            CURRENT_LIA_PREFIXES  = {"2-1", "2-2", "2-3", "2-9"}
            LTLIA_PREFIXES        = {"2-4"}

            EQ_GROUPS = {
                "3-1": "Modal Disetor",
                "3-2": "Tambahan Modal Disetor",
                "3-3": "Laba Ditahan",
            }

            def _prefix(code: str) -> str:
                parts = code.split("-")
                return f"{parts[0]}-{parts[1][:1]}" if len(parts) >= 2 else code[:3]

            # Kumpulkan per tipe
            assets_current   = {}
            assets_fixed     = {}
            liabilities_cur  = {}
            liabilities_lt   = {}
            equity_accts     = {}

            for code, acct in balances.items():
                t = acct["account_type"]
                if t == "asset":
                    p = _prefix(code)
                    grp = ASSET_GROUPS.get(p, "Aset Lainnya")
                    if p in CURRENT_ASSET_PREFIXES:
                        assets_current.setdefault(grp, []).append(acct)
                    else:
                        assets_fixed.setdefault(grp, []).append(acct)
                elif t == "liability":
                    p = _prefix(code)
                    grp = LIA_GROUPS.get(p, "Kewajiban Lainnya")
                    if p in CURRENT_LIA_PREFIXES:
                        liabilities_cur.setdefault(grp, []).append(acct)
                    else:
                        liabilities_lt.setdefault(grp, []).append(acct)
                elif t == "equity":
                    p = _prefix(code)
                    grp = EQ_GROUPS.get(p, "Ekuitas Lainnya")
                    equity_accts.setdefault(grp, []).append(acct)

            def _group_section(groups_dict: dict) -> list:
                result = []
                for grp_name, accts in groups_dict.items():
                    items = [{"account_code": a["account_code"],
                              "account_name": a["account_name"],
                              "balance": float(a["balance"])} for a in accts]
                    subtotal = sum(a["balance"] for a in accts)
                    result.append({"group_name": grp_name, "items": items,
                                   "subtotal": float(subtotal)})
                return result

            curr_assets  = _group_section(assets_current)
            fixed_assets = _group_section(assets_fixed)
            cur_lia      = _group_section(liabilities_cur)
            lt_lia       = _group_section(liabilities_lt)
            eq_groups    = _group_section(equity_accts)

            # Totals
            total_curr_asset  = sum(i["subtotal"] for i in curr_assets)
            total_fixed_asset = sum(i["subtotal"] for i in fixed_assets)
            total_asset       = total_curr_asset + total_fixed_asset

            total_cur_lia = sum(i["subtotal"] for i in cur_lia)
            total_lt_lia  = sum(i["subtotal"] for i in lt_lia)
            total_lia     = total_cur_lia + total_lt_lia

            total_equity_base = sum(i["subtotal"] for i in eq_groups)
            total_equity = float(total_equity_base) + float(net_income)

            return {
                "aktiva": {
                    "aktiva_lancar": {"groups": curr_assets,  "total": total_curr_asset},
                    "aktiva_tetap":  {"groups": fixed_assets, "total": total_fixed_asset},
                    "total":         total_asset,
                },
                "pasiva": {
                    "kewajiban_lancar": {"groups": cur_lia,  "total": total_cur_lia},
                    "kewajiban_jangka_panjang": {"groups": lt_lia, "total": total_lt_lia},
                    "ekuitas": {
                        "groups": eq_groups,
                        "laba_periode_berjalan": float(net_income),
                        "total": total_equity,
                    },
                    "total": total_lia + total_equity,
                },
                "is_balanced": abs(total_asset - (total_lia + total_equity)) < 2,
                "difference":  total_asset - (total_lia + total_equity),
            }

        current_sections = _build_sections(current, curr_net_income)
        compare_sections = _build_sections(compare, comp_net_income) if compare else None

        entity_name = db.execute(text(
            "SELECT entity_name FROM entity WHERE id = :id"
        ), {"id": entity_id}).scalar()

        return {
            "report_type": "balance_sheet",
            "entity_id":   entity_id,
            "entity_name": entity_name,
            "as_of_date":  str(as_of_date),
            "compare_date": str(compare_date) if compare_date else None,
            "current":     current_sections,
            "compare":     compare_sections,
        }

    # ── 3. PROFIT & LOSS ─────────────────────────────────────────────────────

    @staticmethod
    def get_profit_loss(
        db: Session,
        entity_id: str,
        from_date: date,
        to_date: date,
        compare_from: Optional[date] = None,
        compare_to: Optional[date] = None,
    ) -> dict:
        """
        Laporan Laba Rugi untuk rentang tanggal.
        Mendukung kolom pembanding.
        """
        def _fetch(d_from: date, d_to: date) -> dict:
            rows = db.execute(text("""
                SELECT
                    coa.account_code,
                    coa.account_name,
                    coa.account_type,
                    coa.account_group,
                    coa.normal_balance,
                    COALESCE(SUM(gl.debit_idr), 0)  AS total_dr,
                    COALESCE(SUM(gl.credit_idr), 0) AS total_cr
                FROM chart_of_accounts coa
                LEFT JOIN gl_line gl  ON gl.account_id = coa.id
                LEFT JOIN gl_journal gj ON gj.id = gl.journal_id
                                       AND gj.status = 'posted'
                                       AND gj.entity_id = :eid
                                       AND gj.journal_date BETWEEN :df AND :dt
                WHERE coa.entity_id = :eid
                  AND coa.account_type IN (
                      'revenue','cogs','expense',
                      'other_income','other_expense','tax_expense'
                  )
                  AND coa.is_header = FALSE
                GROUP BY coa.account_code, coa.account_name, coa.account_type,
                         coa.account_group, coa.normal_balance
                ORDER BY coa.account_code
            """), {"eid": entity_id, "df": d_from, "dt": d_to}).fetchall()

            by_type: dict = {}
            for r in rows:
                dr = _d(r.total_dr)
                cr = _d(r.total_cr)
                normal = (r.normal_balance or "credit").lower()
                balance = (cr - dr) if normal == "credit" else (dr - cr)
                by_type.setdefault(r.account_type, []).append({
                    "account_code":  r.account_code,
                    "account_name":  r.account_name,
                    "account_group": r.account_group,
                    "balance":       float(balance),
                })
            return by_type

        def _build_pl(by_type: dict) -> dict:
            def _section(key: str):
                items = by_type.get(key, [])
                # Group by account_group if exists
                groups: dict = {}
                for item in items:
                    grp = item.get("account_group") or "Umum"
                    groups.setdefault(grp, []).append(item)
                result = []
                for grp_name, accts in groups.items():
                    subtotal = sum(a["balance"] for a in accts)
                    result.append({"group_name": grp_name, "items": accts,
                                   "subtotal": subtotal})
                return result, sum(a["balance"] for a in items)

            rev_groups,  total_rev    = _section("revenue")
            cogs_groups, total_cogs   = _section("cogs")
            exp_groups,  total_exp    = _section("expense")
            oi_groups,   total_oi     = _section("other_income")
            oe_groups,   total_oe     = _section("other_expense")
            tax_groups,  total_tax    = _section("tax_expense")

            gross_profit  = total_rev - total_cogs
            operating_profit = gross_profit - total_exp
            ebt           = operating_profit + total_oi - total_oe
            net_income    = ebt - total_tax

            return {
                "pendapatan":      {"groups": rev_groups,  "total": total_rev},
                "hpp":             {"groups": cogs_groups, "total": total_cogs},
                "laba_bruto":      gross_profit,
                "beban_operasi":   {"groups": exp_groups,  "total": total_exp},
                "laba_operasi":    operating_profit,
                "pendapatan_lain": {"groups": oi_groups,   "total": total_oi},
                "beban_lain":      {"groups": oe_groups,   "total": total_oe},
                "laba_sebelum_pajak": ebt,
                "beban_pajak":     {"groups": tax_groups,  "total": total_tax},
                "laba_bersih":     net_income,
            }

        current_data    = _fetch(from_date, to_date)
        current_pl      = _build_pl(current_data)
        compare_pl      = _build_pl(_fetch(compare_from, compare_to)) if compare_from else None

        entity_name = db.execute(text(
            "SELECT entity_name FROM entity WHERE id = :id"
        ), {"id": entity_id}).scalar()

        return {
            "report_type":  "profit_loss",
            "entity_id":    entity_id,
            "entity_name":  entity_name,
            "from_date":    str(from_date),
            "to_date":      str(to_date),
            "compare_from": str(compare_from) if compare_from else None,
            "compare_to":   str(compare_to) if compare_to else None,
            "current":      current_pl,
            "compare":      compare_pl,
        }

    # ── 4. GENERAL LEDGER DETAIL ─────────────────────────────────────────────

    @staticmethod
    def get_general_ledger(
        db: Session,
        entity_id: str,
        from_date: date,
        to_date: date,
        account_code: Optional[str] = None,
        account_type: Optional[str] = None,
        journal_type: Optional[str] = None,
        page: int = 1,
        size: int = 100,
    ) -> dict:
        """
        Buku Besar Detail per akun.
        Menampilkan opening balance, setiap transaksi, running balance.
        """
        # Cari akun yang sesuai filter
        acct_conditions = ["coa.entity_id = :eid", "coa.is_header = FALSE"]
        acct_params: dict = {"eid": entity_id}
        if account_code:
            acct_conditions.append("coa.account_code = :code")
            acct_params["code"] = account_code
        if account_type:
            acct_conditions.append("coa.account_type = :atype")
            acct_params["atype"] = account_type

        acct_where = " AND ".join(acct_conditions)
        accounts = db.execute(text(f"""
            SELECT id, account_code, account_name, account_type, normal_balance
            FROM chart_of_accounts coa
            WHERE {acct_where}
            ORDER BY account_code
        """), acct_params).fetchall()

        ledger_accounts = []

        for acct in accounts:
            acct_id = str(acct.id)
            normal = (acct.normal_balance or "debit").lower()

            # Opening balance (semua transaksi sebelum from_date)
            open_row = db.execute(text("""
                SELECT
                    COALESCE(SUM(gl.debit_idr), 0)  AS total_dr,
                    COALESCE(SUM(gl.credit_idr), 0) AS total_cr
                FROM gl_line gl
                JOIN gl_journal gj ON gj.id = gl.journal_id
                                   AND gj.status = 'posted'
                                   AND gj.entity_id = :eid
                                   AND gj.journal_date < :df
                WHERE gl.account_id = :acct_id
            """), {"eid": entity_id, "df": from_date, "acct_id": acct_id}).first()

            open_dr = _d(open_row.total_dr)
            open_cr = _d(open_row.total_cr)
            opening_balance = (open_dr - open_cr) if normal == "debit" else (open_cr - open_dr)

            # Transaksi dalam periode
            jtype_filter = "AND gj.journal_type = :jtype" if journal_type else ""
            jtype_param = {"jtype": journal_type} if journal_type else {}

            trans_rows = db.execute(text(f"""
                SELECT
                    gj.journal_date,
                    gj.journal_number,
                    gj.journal_type,
                    gj.description AS journal_desc,
                    gl.description AS line_desc,
                    gl.debit_idr,
                    gl.credit_idr,
                    gl.currency,
                    gl.amount_fcy,
                    gl.exchange_rate
                FROM gl_line gl
                JOIN gl_journal gj ON gj.id = gl.journal_id
                                   AND gj.status = 'posted'
                                   AND gj.entity_id = :eid
                                   AND gj.journal_date BETWEEN :df AND :dt
                                   {jtype_filter}
                WHERE gl.account_id = :acct_id
                ORDER BY gj.journal_date, gj.journal_number
                LIMIT :size OFFSET :offset
            """), {
                "eid": entity_id, "df": from_date, "dt": to_date,
                "acct_id": acct_id,
                "size": size, "offset": (page - 1) * size,
                **jtype_param,
            }).fetchall()

            # Hitung running balance
            running = opening_balance
            transactions = []
            period_dr = period_cr = Decimal("0")

            for t in trans_rows:
                dr = _d(t.debit_idr)
                cr = _d(t.credit_idr)
                period_dr += dr
                period_cr += cr
                if normal == "debit":
                    running += dr - cr
                else:
                    running += cr - dr

                transactions.append({
                    "date":          str(t.journal_date),
                    "journal_number": t.journal_number,
                    "journal_type":  t.journal_type,
                    "description":   t.line_desc or t.journal_desc,
                    "debit":         float(dr),
                    "credit":        float(cr),
                    "balance":       float(running),
                    "currency":      t.currency,
                    "amount_fcy":    float(_d(t.amount_fcy)) if t.amount_fcy else None,
                })

            closing_balance = (opening_balance +
                               (period_dr - period_cr if normal == "debit"
                                else period_cr - period_dr))

            if not transactions and account_code is None:
                continue  # skip akun tanpa transaksi jika all-accounts mode

            ledger_accounts.append({
                "account_code":    acct.account_code,
                "account_name":    acct.account_name,
                "account_type":    acct.account_type,
                "opening_balance": float(opening_balance),
                "period_debit":    float(period_dr),
                "period_credit":   float(period_cr),
                "closing_balance": float(closing_balance),
                "transactions":    transactions,
            })

        entity_name = db.execute(text(
            "SELECT entity_name FROM entity WHERE id = :id"
        ), {"id": entity_id}).scalar()

        return {
            "report_type": "general_ledger",
            "entity_id":   entity_id,
            "entity_name": entity_name,
            "from_date":   str(from_date),
            "to_date":     str(to_date),
            "accounts":    ledger_accounts,
            "account_count": len(ledger_accounts),
        }

    # ── 5. CASH FLOW STATEMENT (Indirect Method) ─────────────────────────────

    @staticmethod
    def get_cash_flow(
        db: Session,
        entity_id: str,
        from_date: date,
        to_date: date,
        compare_from: Optional[date] = None,
        compare_to: Optional[date] = None,
    ) -> dict:
        """
        Laporan Arus Kas Metode Tidak Langsung (PSAK 2 / IAS 7).

        I.  Arus Kas Operasi
              Laba Bersih
            + Penyesuaian non-kas (depresiasi, amortisasi)
            ± Perubahan modal kerja (ΔAR, ΔInventory, ΔAP, Δakrual)

        II. Arus Kas Investasi
              Pembelian Aset Tetap
              Penjualan Aset Tetap

        III.Arus Kas Pendanaan
              Penerimaan Pinjaman
              Pembayaran Pinjaman
              Setoran Modal
              Pembayaran Dividen

        Klasifikasi akun ke aktivitas berdasarkan account_code prefix:
          1-5xxx, 1-7xxx, 1-8xxx → Investasi
          2-3xxx, 2-4xxx, 3-xxx  → Pendanaan
          Sisanya → Operasi
        """
        def _compute(d_from: date, d_to: date) -> dict:
            # Ambil net income dari P&L
            pl = FinancialReportEngine.get_profit_loss(db, entity_id, d_from, d_to)
            net_income = _d(pl["current"]["laba_bersih"])

            # Ambil semua akun non-P&L dengan perubahan saldo dalam periode
            rows = db.execute(text("""
                SELECT
                    coa.account_code,
                    coa.account_name,
                    coa.account_type,
                    coa.normal_balance,
                    -- Opening balance (sebelum periode)
                    COALESCE(SUM(gl.debit_idr)  FILTER (WHERE gj.journal_date < :df), 0) -
                    COALESCE(SUM(gl.credit_idr) FILTER (WHERE gj.journal_date < :df), 0) AS open_net,
                    -- Closing balance (sampai akhir periode)
                    COALESCE(SUM(gl.debit_idr)  FILTER (WHERE gj.journal_date <= :dt), 0) -
                    COALESCE(SUM(gl.credit_idr) FILTER (WHERE gj.journal_date <= :dt), 0) AS close_net
                FROM chart_of_accounts coa
                LEFT JOIN gl_line gl  ON gl.account_id = coa.id
                LEFT JOIN gl_journal gj ON gj.id = gl.journal_id
                                       AND gj.status = 'posted'
                                       AND gj.entity_id = :eid
                WHERE coa.entity_id = :eid
                  AND coa.account_type IN ('asset','liability','equity')
                  AND coa.is_header = FALSE
                GROUP BY coa.account_code, coa.account_name,
                         coa.account_type, coa.normal_balance
            """), {"eid": entity_id, "df": d_from, "dt": d_to}).fetchall()

            # Klasifikasi
            INVESTING_PREFIXES  = {"1-5", "1-6", "1-7", "1-8"}
            FINANCING_PREFIXES  = {"2-3", "2-4", "3-1", "3-2", "3-3"}
            CASH_PREFIXES       = {"1-1"}                           # Kas & Bank
            AR_PREFIXES         = {"1-2"}
            INVENTORY_PREFIXES  = {"1-3", "1-4"}
            AP_PREFIXES         = {"2-1", "2-2"}

            def _prefix(code: str) -> str:
                parts = code.split("-")
                return f"{parts[0]}-{parts[1][:1]}" if len(parts) >= 2 else code[:3]

            # Depresiasi: perubahan akum. depresiasi (1-6xxx, kredit-normal)
            depreciation = Decimal("0")
            operating_adj = []
            investing     = []
            financing     = []

            cash_open = cash_close = Decimal("0")

            for r in rows:
                code = r.account_code
                pre  = _prefix(code)
                normal = (r.normal_balance or "debit").lower()
                open_net  = _d(r.open_net)
                close_net = _d(r.close_net)

                # Untuk akun credit-normal (liability, equity): flip sign
                if normal == "credit":
                    open_net  = -open_net
                    close_net = -close_net

                change = close_net - open_net  # positif = naik

                if pre in CASH_PREFIXES:
                    cash_open  += open_net
                    cash_close += close_net
                    continue

                if pre in INVESTING_PREFIXES:
                    # Akum. depresiasi (1-6, credit-normal) → add-back non-cash
                    if code.startswith("1-6"):
                        depreciation += abs(change)
                    else:
                        investing.append({
                            "account_code": code,
                            "account_name": r.account_name,
                            "amount": float(-change),  # naik asset = kas keluar
                        })
                elif pre in FINANCING_PREFIXES:
                    if normal == "credit":
                        # Hutang/Ekuitas credit-normal: naik = kas masuk
                        amt = close_net - open_net  # already flipped
                    else:
                        amt = -(close_net - open_net)
                    financing.append({
                        "account_code": code,
                        "account_name": r.account_name,
                        "amount": float(amt),
                    })
                else:
                    # Operasi: perubahan modal kerja
                    if r.account_type == "asset":
                        # Naik asset = pakai kas (negatif utk arus kas)
                        amount = -change
                    else:
                        # Naik liability = tambah kas (positif)
                        amount = change  # already flipped for credit-normal
                    operating_adj.append({
                        "account_code": code,
                        "account_name": r.account_name,
                        "amount":       float(amount),
                    })

            # Hitung total
            total_operating = (float(net_income) + float(depreciation) +
                               sum(a["amount"] for a in operating_adj))
            total_investing  = sum(a["amount"] for a in investing)
            total_financing  = sum(a["amount"] for a in financing)
            net_change_cash  = total_operating + total_investing + total_financing

            return {
                "operasi": {
                    "laba_bersih":    float(net_income),
                    "depresiasi":     float(depreciation),
                    "perubahan_modal_kerja": operating_adj,
                    "total":          total_operating,
                },
                "investasi": {
                    "items": investing,
                    "total": total_investing,
                },
                "pendanaan": {
                    "items": financing,
                    "total": total_financing,
                },
                "saldo_kas_awal":   float(cash_open),
                "kenaikan_kas":     net_change_cash,
                "saldo_kas_akhir":  float(cash_close),
                "selisih_check":    float(cash_close - cash_open - Decimal(str(net_change_cash))),
            }

        current_cf = _compute(from_date, to_date)
        compare_cf = _compute(compare_from, compare_to) if compare_from else None

        entity_name = db.execute(text(
            "SELECT entity_name FROM entity WHERE id = :id"
        ), {"id": entity_id}).scalar()

        return {
            "report_type":  "cash_flow",
            "entity_id":    entity_id,
            "entity_name":  entity_name,
            "from_date":    str(from_date),
            "to_date":      str(to_date),
            "compare_from": str(compare_from) if compare_from else None,
            "compare_to":   str(compare_to) if compare_to else None,
            "current":      current_cf,
            "compare":      compare_cf,
        }

    # ── Export PDF ────────────────────────────────────────────────────────────

    @staticmethod
    def export_pdf(report_data: dict, entity_name: str = "") -> bytes:
        """Generate PDF dari data laporan menggunakan WeasyPrint/xhtml2pdf."""
        html = FinancialReportEngine._render_report_html(report_data, entity_name)
        # Import helper dari invoice_template_engine
        from .invoice_template_engine import _html_to_pdf
        return _html_to_pdf(html)

    @staticmethod
    def _render_report_html(data: dict, entity_name: str) -> str:
        """Render laporan ke HTML yang siap di-print/PDF."""
        rtype = data.get("report_type", "")
        name  = data.get("entity_name") or entity_name

        CSS = """
        <style>
          @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
          * { margin:0; padding:0; box-sizing:border-box; }
          body { font-family: Arial, sans-serif; font-size: 11px; color: #1f2937; padding: 16mm; }
          h1 { font-size: 16px; font-weight: 700; color: #1a56db; }
          h2 { font-size: 11px; font-weight: 600; color: #374151; text-transform: uppercase;
               letter-spacing: 0.5px; margin: 12px 0 4px; border-bottom: 1px solid #e5e7eb; padding-bottom: 2px; }
          .header { text-align: center; margin-bottom: 16px; border-bottom: 2px solid #1a56db; padding-bottom: 10px; }
          .header p { font-size: 11px; color: #6b7280; margin-top: 3px; }
          table { width: 100%; border-collapse: collapse; margin-bottom: 8px; }
          th { background: #1a56db; color: white; padding: 6px 8px; font-size: 10px;
               font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; text-align: left; }
          th.right, td.right { text-align: right; }
          td { padding: 5px 8px; border-bottom: 1px solid #f3f4f6; font-size: 11px; }
          tr.group-header td { background: #f3f4f6; font-weight: 600; font-size: 10px; color: #374151; }
          tr.subtotal td { font-weight: 600; border-top: 1px solid #d1d5db; }
          tr.total td { font-weight: 700; font-size: 12px; background: #eff6ff; border-top: 2px solid #1a56db; }
          tr.net-income td { font-weight: 700; font-size: 13px; background: #1a56db; color: white; }
          .positive { color: #059669; }
          .negative { color: #dc2626; }
          .indent { padding-left: 20px; }
          .footer { text-align: center; font-size: 10px; color: #9ca3af;
                    border-top: 1px solid #e5e7eb; padding-top: 8px; margin-top: 16px; }
          @media print { body { padding: 10mm; } }
        </style>
        """

        if rtype == "trial_balance":
            body = _render_trial_balance_html(data)
        elif rtype == "balance_sheet":
            body = _render_balance_sheet_html(data)
        elif rtype == "profit_loss":
            body = _render_pl_html(data)
        elif rtype == "general_ledger":
            body = _render_gl_html(data)
        elif rtype == "cash_flow":
            body = _render_cashflow_html(data)
        else:
            body = f"<p>Report type '{rtype}' tidak dikenal.</p>"

        from datetime import datetime
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">{CSS}</head>
<body>{body}
<div class="footer">Dicetak: {now} &nbsp;|&nbsp; {name}</div>
</body></html>"""

    # ── Export Excel ──────────────────────────────────────────────────────────

    @staticmethod
    def export_excel(report_data: dict) -> bytes:
        """Export laporan ke Excel (.xlsx) menggunakan openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise RuntimeError("openpyxl belum terinstall. Jalankan: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        rtype = report_data.get("report_type", "report")
        ws.title = rtype[:31]

        blue_fill = PatternFill("solid", fgColor="1A56DB")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        bold_font   = Font(name="Arial", bold=True, size=10)
        normal_font = Font(name="Arial", size=10)
        total_fill  = PatternFill("solid", fgColor="EFF6FF")
        total_font  = Font(name="Arial", bold=True, size=11)

        thin = Side(style="thin")
        thick = Side(style="medium")

        def _header_row(ws, values: list, row: int):
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = blue_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        def _data_row(ws, values: list, row: int, bold=False, indent=0, fill=None):
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = bold_font if bold else normal_font
                if fill:
                    cell.fill = fill
                if col == 1 and indent:
                    cell.alignment = Alignment(indent=indent)
                if isinstance(val, (int, float)) and col > 1:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")

        row = 1
        ws.cell(row=row, column=1, value=report_data.get("entity_name", "")).font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value=rtype.replace("_", " ").title()).font = Font(size=12)
        row += 2

        if rtype == "trial_balance":
            _excel_trial_balance(ws, report_data, row, _header_row, _data_row, bold_font, total_fill, total_font)
        elif rtype == "profit_loss":
            _excel_pl(ws, report_data, row, _header_row, _data_row, bold_font, total_fill, total_font)
        elif rtype == "balance_sheet":
            _excel_bs(ws, report_data, row, _header_row, _data_row, bold_font, total_fill, total_font)
        elif rtype == "cash_flow":
            _excel_cf(ws, report_data, row, _header_row, _data_row, bold_font, total_fill, total_font)
        elif rtype == "general_ledger":
            _excel_gl(ws, report_data, row, _header_row, _data_row, bold_font, total_fill, total_font)

        # Auto-width kolom
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ── HTML render helpers ───────────────────────────────────────────────────────

def _render_trial_balance_html(data: dict) -> str:
    t = data["totals"]
    rows_html = ""
    for a in data["accounts"]:
        rows_html += f"""<tr>
          <td>{a['account_code']}</td>
          <td>{a['account_name']}</td>
          <td class="right">{_fmt(a['opening_debit']) if a['opening_debit'] else '-'}</td>
          <td class="right">{_fmt(a['opening_credit']) if a['opening_credit'] else '-'}</td>
          <td class="right">{_fmt(a['period_debit']) if a['period_debit'] else '-'}</td>
          <td class="right">{_fmt(a['period_credit']) if a['period_credit'] else '-'}</td>
          <td class="right">{_fmt(a['closing_debit']) if a['closing_debit'] else '-'}</td>
          <td class="right">{_fmt(a['closing_credit']) if a['closing_credit'] else '-'}</td>
        </tr>"""

    return f"""
    <div class="header">
      <h1>{data.get('entity_name','')}</h1>
      <p>NERACA SALDO (TRIAL BALANCE)<br>
         Periode: {data['fiscal_year']}/{data['fiscal_month']:02d} &nbsp;|&nbsp; Per: {data['as_of_date']}</p>
    </div>
    <table>
      <thead><tr>
        <th>Kode</th><th>Nama Akun</th>
        <th class="right">Saldo Awal Dr</th><th class="right">Saldo Awal Cr</th>
        <th class="right">Mutasi Dr</th><th class="right">Mutasi Cr</th>
        <th class="right">Saldo Akhir Dr</th><th class="right">Saldo Akhir Cr</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
      <tfoot><tr class="total">
        <td colspan="2">TOTAL</td>
        <td class="right">{_fmt(t['opening_debit'])}</td>
        <td class="right">{_fmt(t['opening_credit'])}</td>
        <td class="right">{_fmt(t['period_debit'])}</td>
        <td class="right">{_fmt(t['period_credit'])}</td>
        <td class="right">{_fmt(t['closing_debit'])}</td>
        <td class="right">{_fmt(t['closing_credit'])}</td>
      </tr></tfoot>
    </table>
    <p style="font-size:10px;color:{'#059669' if t['is_balanced'] else '#dc2626'}">
      {'✓ Neraca saldo balance' if t['is_balanced'] else f'⚠ Selisih: Rp {_fmt(t["difference"])}'}
    </p>"""


def _render_balance_sheet_html(data: dict) -> str:
    c = data["current"]

    def _groups(groups_list, indent=True):
        html = ""
        for grp in groups_list:
            html += f'<tr class="group-header"><td colspan="2" class="indent">{grp["group_name"]}</td></tr>'
            for item in grp["items"]:
                html += f'<tr><td class="indent" style="padding-left:24px">{item["account_code"]} {item["account_name"]}</td><td class="right">{_fmt(item["balance"])}</td></tr>'
            html += f'<tr class="subtotal"><td class="indent">Jumlah {grp["group_name"]}</td><td class="right">{_fmt(grp["subtotal"])}</td></tr>'
        return html

    return f"""
    <div class="header">
      <h1>{data.get('entity_name','')}</h1>
      <p>NERACA (BALANCE SHEET)<br>Per Tanggal: {data['as_of_date']}</p>
    </div>
    <table>
      <thead><tr><th>AKUN</th><th class="right">JUMLAH (Rp)</th></tr></thead>
      <tbody>
        <tr class="group-header"><td colspan="2">AKTIVA LANCAR</td></tr>
        {_groups(c['aktiva']['aktiva_lancar']['groups'])}
        <tr class="subtotal"><td>TOTAL AKTIVA LANCAR</td><td class="right">{_fmt(c['aktiva']['aktiva_lancar']['total'])}</td></tr>
        <tr class="group-header"><td colspan="2">AKTIVA TETAP</td></tr>
        {_groups(c['aktiva']['aktiva_tetap']['groups'])}
        <tr class="subtotal"><td>TOTAL AKTIVA TETAP</td><td class="right">{_fmt(c['aktiva']['aktiva_tetap']['total'])}</td></tr>
        <tr class="total"><td>TOTAL AKTIVA</td><td class="right">{_fmt(c['aktiva']['total'])}</td></tr>

        <tr style="height:8px"><td colspan="2"></td></tr>

        <tr class="group-header"><td colspan="2">KEWAJIBAN LANCAR</td></tr>
        {_groups(c['pasiva']['kewajiban_lancar']['groups'])}
        <tr class="subtotal"><td>TOTAL KEWAJIBAN LANCAR</td><td class="right">{_fmt(c['pasiva']['kewajiban_lancar']['total'])}</td></tr>
        <tr class="group-header"><td colspan="2">KEWAJIBAN JANGKA PANJANG</td></tr>
        {_groups(c['pasiva']['kewajiban_jangka_panjang']['groups'])}
        <tr class="subtotal"><td>TOTAL KEWAJIBAN JANGKA PANJANG</td><td class="right">{_fmt(c['pasiva']['kewajiban_jangka_panjang']['total'])}</td></tr>

        <tr class="group-header"><td colspan="2">EKUITAS</td></tr>
        {_groups(c['pasiva']['ekuitas']['groups'])}
        <tr><td class="indent" style="padding-left:24px">Laba Periode Berjalan</td>
            <td class="right">{_fmt(c['pasiva']['ekuitas']['laba_periode_berjalan'])}</td></tr>
        <tr class="subtotal"><td>TOTAL EKUITAS</td><td class="right">{_fmt(c['pasiva']['ekuitas']['total'])}</td></tr>

        <tr class="total"><td>TOTAL KEWAJIBAN DAN EKUITAS</td><td class="right">{_fmt(c['pasiva']['total'])}</td></tr>
      </tbody>
    </table>
    <p style="font-size:10px;color:{'#059669' if c['is_balanced'] else '#dc2626'}">
      {'✓ Neraca balance' if c['is_balanced'] else f'⚠ Selisih: Rp {_fmt(c["difference"])}'}
    </p>"""


def _render_pl_html(data: dict) -> str:
    c = data["current"]

    def _section(sec_data):
        html = ""
        for grp in sec_data.get("groups", []):
            html += f'<tr class="group-header"><td class="indent">{grp["group_name"]}</td><td></td></tr>'
            for item in grp["items"]:
                html += f'<tr><td class="indent" style="padding-left:24px">{item["account_code"]} {item["account_name"]}</td><td class="right">{_fmt(item["balance"])}</td></tr>'
        return html

    return f"""
    <div class="header">
      <h1>{data.get('entity_name','')}</h1>
      <p>LAPORAN LABA RUGI<br>{data['from_date']} s/d {data['to_date']}</p>
    </div>
    <table>
      <thead><tr><th>KETERANGAN</th><th class="right">JUMLAH (Rp)</th></tr></thead>
      <tbody>
        <tr class="group-header"><td colspan="2">PENDAPATAN</td></tr>
        {_section(c['pendapatan'])}
        <tr class="subtotal"><td>TOTAL PENDAPATAN</td><td class="right">{_fmt(c['pendapatan']['total'])}</td></tr>
        <tr class="group-header"><td colspan="2">HARGA POKOK PENJUALAN</td></tr>
        {_section(c['hpp'])}
        <tr class="subtotal"><td>TOTAL HPP</td><td class="right">({_fmt(c['hpp']['total'])})</td></tr>
        <tr class="total"><td>LABA BRUTO</td><td class="right">{_fmt(c['laba_bruto'])}</td></tr>
        <tr class="group-header"><td colspan="2">BEBAN OPERASIONAL</td></tr>
        {_section(c['beban_operasi'])}
        <tr class="subtotal"><td>TOTAL BEBAN OPERASIONAL</td><td class="right">({_fmt(c['beban_operasi']['total'])})</td></tr>
        <tr class="total"><td>LABA OPERASIONAL (EBIT)</td><td class="right">{_fmt(c['laba_operasi'])}</td></tr>
        {_section(c['pendapatan_lain'])}
        {_section(c['beban_lain'])}
        <tr class="total"><td>LABA SEBELUM PAJAK</td><td class="right">{_fmt(c['laba_sebelum_pajak'])}</td></tr>
        {_section(c['beban_pajak'])}
        <tr class="net-income"><td>LABA BERSIH</td><td class="right">{_fmt(c['laba_bersih'])}</td></tr>
      </tbody>
    </table>"""


def _render_gl_html(data: dict) -> str:
    body = f"""
    <div class="header">
      <h1>{data.get('entity_name','')}</h1>
      <p>BUKU BESAR (GENERAL LEDGER)<br>{data['from_date']} s/d {data['to_date']}</p>
    </div>"""

    for acct in data["accounts"]:
        rows_html = ""
        for t in acct["transactions"]:
            rows_html += f"""<tr>
              <td>{t['date']}</td>
              <td>{t['journal_number']}</td>
              <td>{t['description'] or ''}</td>
              <td class="right">{_fmt(t['debit']) if t['debit'] else '-'}</td>
              <td class="right">{_fmt(t['credit']) if t['credit'] else '-'}</td>
              <td class="right">{_fmt(t['balance'])}</td>
            </tr>"""

        body += f"""
        <h2>{acct['account_code']} — {acct['account_name']}</h2>
        <table>
          <thead><tr>
            <th>Tanggal</th><th>No. Jurnal</th><th>Keterangan</th>
            <th class="right">Debit</th><th class="right">Kredit</th><th class="right">Saldo</th>
          </tr></thead>
          <tbody>
            <tr class="group-header">
              <td colspan="3">Saldo Awal</td>
              <td class="right"></td><td class="right"></td>
              <td class="right">{_fmt(acct['opening_balance'])}</td>
            </tr>
            {rows_html}
          </tbody>
          <tfoot><tr class="total">
            <td colspan="3">SALDO AKHIR</td>
            <td class="right">{_fmt(acct['period_debit'])}</td>
            <td class="right">{_fmt(acct['period_credit'])}</td>
            <td class="right">{_fmt(acct['closing_balance'])}</td>
          </tr></tfoot>
        </table>"""
    return body


def _render_cashflow_html(data: dict) -> str:
    c = data["current"]

    def _items(items_list):
        html = ""
        for item in items_list:
            html += f'<tr><td class="indent" style="padding-left:20px">{item["account_name"]}</td><td class="right">{_fmt(item["amount"])}</td></tr>'
        return html

    return f"""
    <div class="header">
      <h1>{data.get('entity_name','')}</h1>
      <p>LAPORAN ARUS KAS (METODE TIDAK LANGSUNG)<br>{data['from_date']} s/d {data['to_date']}</p>
    </div>
    <table>
      <thead><tr><th>KETERANGAN</th><th class="right">JUMLAH (Rp)</th></tr></thead>
      <tbody>
        <tr class="group-header"><td colspan="2">I. ARUS KAS DARI AKTIVITAS OPERASI</td></tr>
        <tr><td class="indent">Laba Bersih</td><td class="right">{_fmt(c['operasi']['laba_bersih'])}</td></tr>
        <tr><td class="indent">Penyusutan dan Amortisasi</td><td class="right">{_fmt(c['operasi']['depresiasi'])}</td></tr>
        <tr class="group-header"><td class="indent">Perubahan Modal Kerja:</td><td></td></tr>
        {_items(c['operasi']['perubahan_modal_kerja'])}
        <tr class="total"><td>TOTAL ARUS KAS OPERASI</td><td class="right">{_fmt(c['operasi']['total'])}</td></tr>

        <tr style="height:6px"><td colspan="2"></td></tr>
        <tr class="group-header"><td colspan="2">II. ARUS KAS DARI AKTIVITAS INVESTASI</td></tr>
        {_items(c['investasi']['items'])}
        <tr class="total"><td>TOTAL ARUS KAS INVESTASI</td><td class="right">{_fmt(c['investasi']['total'])}</td></tr>

        <tr style="height:6px"><td colspan="2"></td></tr>
        <tr class="group-header"><td colspan="2">III. ARUS KAS DARI AKTIVITAS PENDANAAN</td></tr>
        {_items(c['pendanaan']['items'])}
        <tr class="total"><td>TOTAL ARUS KAS PENDANAAN</td><td class="right">{_fmt(c['pendanaan']['total'])}</td></tr>

        <tr style="height:8px"><td colspan="2"></td></tr>
        <tr class="subtotal"><td>Saldo Kas Awal Periode</td><td class="right">{_fmt(c['saldo_kas_awal'])}</td></tr>
        <tr class="subtotal"><td>Kenaikan / (Penurunan) Kas Bersih</td><td class="right">{_fmt(c['kenaikan_kas'])}</td></tr>
        <tr class="net-income"><td>SALDO KAS AKHIR PERIODE</td><td class="right">{_fmt(c['saldo_kas_akhir'])}</td></tr>
      </tbody>
    </table>"""


# ── Excel helpers (minimal stubs) ─────────────────────────────────────────────

def _excel_trial_balance(ws, data, row, hdr, drow, bold_font, total_fill, total_font):
    hdr(ws, ["Kode Akun","Nama Akun","Saldo Awal Dr","Saldo Awal Cr",
             "Mutasi Dr","Mutasi Cr","Saldo Akhir Dr","Saldo Akhir Cr"], row); row += 1
    for a in data["accounts"]:
        drow(ws, [a["account_code"], a["account_name"],
                  a["opening_debit"], a["opening_credit"],
                  a["period_debit"],  a["period_credit"],
                  a["closing_debit"], a["closing_credit"]], row)
        row += 1
    t = data["totals"]
    drow(ws, ["","TOTAL", t["opening_debit"], t["opening_credit"],
              t["period_debit"], t["period_credit"],
              t["closing_debit"], t["closing_credit"]], row, bold=True, fill=total_fill)


def _excel_pl(ws, data, row, hdr, drow, bold_font, total_fill, total_font):
    c = data["current"]
    hdr(ws, ["Keterangan", "Jumlah (Rp)",
             "Perbandingan (Rp)" if data.get("compare") else ""], row); row += 1
    comp = data.get("compare")

    def _section(label, sec, comp_sec=None):
        nonlocal row
        drow(ws, [label], row, bold=True); row += 1
        for grp in sec.get("groups", []):
            for item in grp["items"]:
                comp_val = None
                if comp_sec:
                    for cg in comp_sec.get("groups", []):
                        for ci in cg["items"]:
                            if ci["account_code"] == item["account_code"]:
                                comp_val = ci["balance"]
                drow(ws, ["  " + item["account_code"] + " " + item["account_name"],
                          item["balance"], comp_val], row)
                row += 1
        drow(ws, ["TOTAL " + label, sec["total"],
                  comp_sec["total"] if comp_sec else None], row, bold=True, fill=total_fill)
        row += 1

    _section("PENDAPATAN",       c["pendapatan"],    comp["pendapatan"] if comp else None)
    _section("HPP",              c["hpp"],           comp["hpp"] if comp else None)
    drow(ws, ["LABA BRUTO", c["laba_bruto"]], row, bold=True, fill=total_fill); row += 1
    _section("BEBAN OPERASI",    c["beban_operasi"], comp["beban_operasi"] if comp else None)
    drow(ws, ["LABA OPERASI", c["laba_operasi"]], row, bold=True, fill=total_fill); row += 1
    drow(ws, ["LABA BERSIH", c["laba_bersih"]], row, bold=True, fill=total_fill); row += 1


def _excel_bs(ws, data, row, hdr, drow, bold_font, total_fill, total_font):
    c = data["current"]
    hdr(ws, ["Akun", "Jumlah (Rp)"], row); row += 1
    drow(ws, ["AKTIVA"], row, bold=True); row += 1
    for grp in c["aktiva"]["aktiva_lancar"]["groups"]:
        drow(ws, ["  " + grp["group_name"]], row, bold=True); row += 1
        for item in grp["items"]:
            drow(ws, ["    " + item["account_code"] + " " + item["account_name"], item["balance"]], row)
            row += 1
    drow(ws, ["TOTAL AKTIVA", c["aktiva"]["total"]], row, bold=True, fill=total_fill); row += 1


def _excel_cf(ws, data, row, hdr, drow, bold_font, total_fill, total_font):
    c = data["current"]
    hdr(ws, ["Keterangan", "Jumlah (Rp)"], row); row += 1
    drow(ws, ["I. ARUS KAS OPERASI"], row, bold=True); row += 1
    drow(ws, ["  Laba Bersih", c["operasi"]["laba_bersih"]], row); row += 1
    drow(ws, ["  Penyusutan", c["operasi"]["depresiasi"]], row); row += 1
    for item in c["operasi"]["perubahan_modal_kerja"]:
        drow(ws, ["  " + item["account_name"], item["amount"]], row); row += 1
    drow(ws, ["TOTAL ARUS KAS OPERASI", c["operasi"]["total"]], row, bold=True, fill=total_fill); row += 1
    drow(ws, ["SALDO KAS AKHIR", c["saldo_kas_akhir"]], row, bold=True, fill=total_fill); row += 1


def _excel_gl(ws, data, row, hdr, drow, bold_font, total_fill, total_font):
    hdr(ws, ["Tanggal","No. Jurnal","Keterangan","Debit","Kredit","Saldo"], row); row += 1
    for acct in data["accounts"]:
        drow(ws, [acct["account_code"] + " - " + acct["account_name"]], row, bold=True); row += 1
        for t in acct["transactions"]:
            drow(ws, [t["date"], t["journal_number"], t["description"],
                      t["debit"], t["credit"], t["balance"]], row)
            row += 1
        drow(ws, ["", "", "SALDO AKHIR", acct["period_debit"],
                  acct["period_credit"], acct["closing_balance"]], row, bold=True, fill=total_fill)
        row += 2
